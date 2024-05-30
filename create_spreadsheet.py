######################################################################
# Author: Sam McFarland
#
# Purpose: A program which takes user input and writes it to an Excel file to keep track of basic travel itinerary.
######################################################################
import openpyxl
from openpyxl.styles import Font


def create_excel_file():
    """
    Takes user input and creates an Excel file filled with the information given.
    :return: None
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the headers in the first row
    headers = ["Officer", "Location", "Dates", "Donor Name"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=14)  # Make the text bold and larger

    # Get user input and add to the worksheet
    while True:
        officer = input("Enter Officer name (or type 'done' to finish): ")
        if officer.lower() == 'done':
            break
        location = input("Enter Location: ")
        dates = input("Enter Dates: ")
        donor_name = input("Enter Donor Name: ")

        # Find the next available row in the sheet
        next_row = sheet.max_row + 1

        # Write the user input to the next row
        sheet.cell(row=next_row, column=1, value=officer)
        sheet.cell(row=next_row, column=2, value=location)
        sheet.cell(row=next_row, column=3, value=dates)
        sheet.cell(row=next_row, column=4, value=donor_name)

    # Save the workbook
    workbook.save("donations.xlsx")
    print("Data saved to donations.xlsx")


def main():
    create_excel_file()


if __name__ == "__main__":
    main()

